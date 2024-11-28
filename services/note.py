from fastapi import HTTPException, status

from database.base_crud import BaseCRUD
from dto.note_dto import NoteCreateRequest, NoteResponse, TagResponse, NoteUpdateRequest
from models.notes import Note, NotesTags
from security import decode_access_token
from database.database import async_session_maker
from sqlalchemy import select
from sqlalchemy.orm import joinedload


class NoteCRUD(BaseCRUD):
    model = Note

    @classmethod
    async def get_all(cls, user_id: int):
        async with async_session_maker() as session:
            query = (
                select(cls.model)
                .options(joinedload(cls.model.tags).joinedload(NotesTags.tag))
                .where(cls.model.user_id == user_id, cls.model.is_archive == False)
                .order_by(cls.model.created_at.desc())
            )
            res = await session.execute(query)
            return res.unique().scalars().all()

    @classmethod
    async def get_all_archive(cls, user_id: int):
        async with async_session_maker() as session:
            query = (
                select(cls.model)
                .options(joinedload(cls.model.tags).joinedload(NotesTags.tag))
                .where(cls.model.user_id == user_id, cls.model.is_archive == True)
                .order_by(cls.model.created_at.desc())
            )
            res = await session.execute(query)
            return res.unique().scalars().all()

    @classmethod
    async def get_one(cls, note_id: int):
        async with async_session_maker() as session:
            query = (
                select(cls.model)
                .options(joinedload(cls.model.tags).joinedload(NotesTags.tag))
                .where(cls.model.id == note_id)
            )
            res = await session.execute(query)
            return res.unique().scalar_one_or_none()


class NoteTagsCRUD(BaseCRUD):
    model = NotesTags


async def get_all_my(token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    notes = await NoteCRUD.get_all(user_id=int(user_id))
    return [
        NoteResponse(
            id=note.id,
            title=note.title,
            user_id=note.user_id,
            description=note.description,
            is_archive=note.is_archive,
            created_at=note.created_at,
            updated_at=note.updated_at,
            tags=[
                TagResponse(
                    id=note_tag.tag.id,
                    name=note_tag.tag.name,
                    color=note_tag.tag.color,
                )
                for note_tag in note.tags
            ]
        )
        for note in notes
    ]

async def get_all_my_archives(token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    notes = await NoteCRUD.get_all_archive(user_id=int(user_id))
    return [
        NoteResponse(
            id=note.id,
            title=note.title,
            user_id=note.user_id,
            description=note.description,
            is_archive=note.is_archive,
            created_at=note.created_at,
            updated_at=note.updated_at,
            tags=[
                TagResponse(
                    id=note_tag.tag.id,
                    name=note_tag.tag.name,
                    color=note_tag.tag.color,
                )
                for note_tag in note.tags
            ]
        )
        for note in notes
    ]

async def get_by_id(note_id: int) -> NoteResponse:
    note = await NoteCRUD.get_one(note_id)
    if not note:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)

    return NoteResponse(
        id=note.id,
        title=note.title,
        user_id=note.user_id,
        description=note.description,
        is_archive=note.is_archive,
        created_at=note.created_at,
        updated_at=note.updated_at,
        tags=[
            TagResponse(
                id=note_tag.tag.id,
                name=note_tag.tag.name,
                color=note_tag.tag.color,
            )
            for note_tag in note.tags
        ]
    )


async def create_note(data: NoteCreateRequest, token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    note_id = await NoteCRUD.create_and_return_id(user_id=int(user_id), title=data.title, description=data.description)

    if data.noteTags:
        for tag_id in data.noteTags:
            await NoteTagsCRUD.create(note_id=note_id, tag_id=tag_id)

    return {"ok": True, "note_id": note_id}


async def delete_by_id(note_id: int, token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    note = await NoteCRUD.find_by_id(model_id=note_id)
    if note is None:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if note.user_id != int(user_id):
        raise HTTPException(status_code=403, detail="Изменения не доступны")

    await NoteCRUD.delete(model_id=note_id)


async def archive_add_by_id(note_id: int, token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    note = await NoteCRUD.find_by_id(model_id=note_id)
    if note is None:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if note.user_id != int(user_id):
        raise HTTPException(status_code=403, detail="Изменения не доступны")

    await NoteCRUD.update(model_id=note_id, is_archive=True)

async def archive_remove_by_id(note_id: int, token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    note = await NoteCRUD.find_by_id(model_id=note_id)
    if note is None:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if note.user_id != int(user_id):
        raise HTTPException(status_code=403, detail="Изменения не доступны")

    await NoteCRUD.update(model_id=note_id, is_archive=False)


async def update_note(note_id: int, data: NoteUpdateRequest, token: str):
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    user_id = payload.get("sub")
    if not isinstance(user_id, str) or not user_id:
        raise HTTPException(status_code=401, detail="Не валидный токен")

    note = await NoteCRUD.get_one(note_id)
    if not note:
        raise HTTPException(status_code=404, detail="Заметка не найдена")
    
    if str(note.user_id) != user_id:
        raise HTTPException(status_code=403, detail="Нет доступа к заметке")

    # Обновляем основные поля заметки
    note.title = data.title
    note.description = data.description

    # Обновляем теги
    await NoteTagsCRUD.delete_many(note_id=note.id)
    for tag_id in data.tags:
        await NoteTagsCRUD.create({"note_id": note.id, "tag_id": tag_id})

    await NoteCRUD.update(note)
    return await get_by_id(note.id)


from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from fastapi.responses import StreamingResponse


async def export_to_excel(token: str):
    # Получаем все заметки пользователя
    notes = await get_all_my(token)
    
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Мои заметки"
    
    # Добавляем заголовки
    headers = ["ID", "Заголовок", "Описание", "Теги", "Дата создания", "Дата обновления"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Добавляем данные
    for row, note in enumerate(notes, 2):
        ws.cell(row=row, column=1, value=note.id)
        ws.cell(row=row, column=2, value=note.title)
        ws.cell(row=row, column=3, value=note.description)
        ws.cell(row=row, column=4, value=", ".join(tag.name for tag in note.tags))
        ws.cell(row=row, column=5, value=note.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=6, value=note.updated_at.strftime("%Y-%m-%d %H:%M:%S") if note.updated_at else "")
    
    # Настраиваем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Формируем имя файла
    filename = f"notes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Возвращаем файл для скачивания
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )